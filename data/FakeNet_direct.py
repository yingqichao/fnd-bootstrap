import random

import cv2
import pickle
import torch
import torch.utils.data as data

import torchvision.transforms.functional as F

from PIL import Image
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from tqdm import tqdm
from transformers import BertModel, BertTokenizer
import clip
import torchvision
import torchvision.transforms as transforms
import torch.nn as nn

class FakeNet_dataset(data.Dataset):

    def __init__(self, root_path='/home/groupshare/AAAI_dataset', dataset='gossip',
                 is_filter=True, is_use_unimodal=True, image_size=224, is_train=True,
                 data_augment=False, with_ambiguity=False, use_soft_label=False,is_sample_positive=1,
                 duplicate_fake_times=1):
        import data.util as util
        self.duplicate_fake_times = duplicate_fake_times
        self.with_ambiguity = with_ambiguity
        self.use_soft_label = use_soft_label
        self.data_augment = data_augment
        self.dataset_name = dataset
        assert (self.dataset_name=='politi' or self.dataset_name=='gossip'), "Error! Only \'gossip\' or \'politi\' supported!"
        super(FakeNet_dataset, self).__init__()
        self.is_sample_positive = is_sample_positive
        print("duplicate_fake_times: {}".format(self.duplicate_fake_times))
        print("is_sample_positive: {}".format(self.is_sample_positive))
        print("Dataset: {}".format(self.dataset_name))
        print("Using More Negative Examples: {}".format(self.data_augment))
        print("Using AMBIGUITY LEARNING: {}".format(self.with_ambiguity))
        print("Using SOFT LABELS: {}".format(self.use_soft_label))
        self.is_train = is_train
        self.root_path = root_path
        self.index = 0
        self.text_max_len = 170
        # self.token = BertTokenizer.from_pretrained('bert-base-uncased')
        self.label_dict, self.label_ambiguity = [], []
        self.image_size = image_size
        self.resize_and_to_tensor = transforms.Compose([
                transforms.Resize((224,224)),
                transforms.ToTensor(),]
        )
        # dataset_names = ['gossip','politi']
        # for dataset_name in dataset_names:
        # wb = openpyxl.load_workbook(root_path+'/{}_{}.xlsx'.format(dataset_name, 'train' if is_train else 'test'))
        wb = openpyxl.load_workbook(root_path + '/{}/{}_{}{}.xlsx'
            .format(self.dataset_name,self.dataset_name, 'train' if is_train else 'test','' if is_filter else '_no_filt'))
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet['C' + str(i)].value)
            label = int(sheet['D' + str(i)].value)
            # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
            label = 1 if label==0 else 0
            if self.is_sample_positive<1 and label==0 and random.random()>self.is_sample_positive:
                continue
            content = str(sheet['B' + str(i)].value)
            category = str(sheet['E' + str(i)].value)
            ##  NEWS TYPE: 0 MULTIMODAL 1 IMAGE 2 TEXT
            if not is_use_unimodal or "multi" in category:
                category = 0
            elif "image" in category:
                category = 1
            else:
                category = 2
            # imgs = images_name.split('|')
            record = {}
            record['images'] = images_name
            record['label'] = label
            record['content'] = content
            record['subfolder'] = '{}_{}'.format(self.dataset_name, 'train' if is_train else 'test')
            record['category'] = category
            self.label_dict.append(record)
            if label == 1 and self.is_train:
                for times in range(self.duplicate_fake_times):
                    self.label_dict.append(record)

        # # DATA AUGMENTATION
        # if self.dataset_name=='gossip' and self.is_train and self.data_augment:
        #     wb = openpyxl.load_workbook(root_path + '/{}/{}_{}.xlsx'
        #                                 .format(self.dataset_name, self.dataset_name, 'more_negative'))
        #     sheetnames = wb.sheetnames
        #     sheet = wb[sheetnames[0]]
        #     rows = sheet.max_row
        #     for i in tqdm(range(2, rows + 1)):
        #         images_name = str(sheet['C' + str(i)].value)
        #         label = int(sheet['D' + str(i)].value)
        #         # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
        #         label = 1 if label == 0 else 0
        #         content = str(sheet['B' + str(i)].value)
        #         category = str(sheet['E' + str(i)].value)
        #         ##  NEWS TYPE: 0 MULTIMODAL 1 IMAGE 2 TEXT
        #         if not is_use_unimodal or "multi" in category:
        #             category = 0
        #         elif "image" in category:
        #             category = 1
        #         else:
        #             category = 2
        #         # imgs = images_name.split('|')
        #         record = {}
        #         record['images'] = images_name
        #         record['label'] = label
        #         record['content'] = content
        #         record['subfolder'] = '{}_{}'.format(self.dataset_name, 'train' if is_train else 'test')
        #         record['category'] = category
        #         self.label_dict.append(record)

        # AMBIGUITY LEARNING
        if self.dataset_name == 'gossip' and self.is_train and self.with_ambiguity:
            wb = openpyxl.load_workbook(root_path + '/{}/{}_{}.xlsx'
                                        .format(self.dataset_name, self.dataset_name, 'train_ambiguity'))
            sheetnames = wb.sheetnames
            sheet = wb[sheetnames[0]]
            rows = sheet.max_row
            for i in tqdm(range(2, rows + 1)):
                images_name = str(sheet['C' + str(i)].value)
                label = int(sheet['D' + str(i)].value)
                # 1 stands for non-related
                content = str(sheet['B' + str(i)].value)
                category = str(sheet['E' + str(i)].value)
                ##  NEWS TYPE: 0 MULTIMODAL 1 IMAGE 2 TEXT
                if not is_use_unimodal or "multi" in category:
                    category = 0
                elif "image" in category:
                    category = 1
                else:
                    category = 2
                # imgs = images_name.split('|')
                record = {}
                record['images'] = images_name
                record['label'] = label
                record['content'] = content
                record['subfolder'] = '{}_{}'.format(self.dataset_name, 'train' if is_train else 'test')
                record['category'] = category
                self.label_ambiguity.append(record)

        # 9:1 division
        # division_idx = int(0.9 * len(self.label_dict))
        # if is_train:
        #     self.label_dict = self.label_dict[:division_idx]
        # else:
        #     self.label_dict = self.label_dict[division_idx:]
        assert len(self.label_dict)!=0, 'Error: GT path is empty.'

        self.pickle_dict = None
        filename = f'{"./train" if is_train else "val"}_{dataset}.pickle'
        print(f"Loading {filename}")
        with open(filename, 'rb') as f:
            self.pickle_dict = pickle.load(f)

    def __getitem__(self, index):

        GT_size = self.image_size # 这个自己指定一下

        # get GT image
        while True:
            record = self.label_dict[index]
            images, label, content = record['images'], record['label'], record['content']
            category = record['category']
            # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来

            GT_path = images #imgs[np.random.randint(0,len(imgs))]
            # if '/' in GT_path:
            #     # excel中文件名前面可能会有'(non)rumor_images/'，这个在读图的时候是不需要的，直接过滤
            #     GT_path = GT_path[GT_path.rfind('/')+1:]
            img_GT = None
            GT_path = "{}/{}/{}/{}".format(self.root_path,'Images', record['subfolder'], GT_path)
            if GT_path in self.pickle_dict and content in self.pickle_dict:
                image_feature = self.pickle_dict[GT_path]
                content_feature = self.pickle_dict[content]
                break
            else:
                print(f"Not found {GT_path}")

        content_feature = torch.from_numpy(content_feature).float()
        image_feature = torch.from_numpy(image_feature).float()

        # AMBIGUITY: SAME, REPEATED CODE
        if self.with_ambiguity:
            # get GT image
            index1 = random.randint(0,len(self.label_ambiguity)-1)
            record = self.label_ambiguity[index1]
            images_ambiguity, label_ambiguity, content_ambiguity = record['images'], record['label'], record['content']
            if self.use_soft_label:
                if label_ambiguity==0:  label_ambiguity=0.2
                else:   label_ambiguity=0.8

            GT_path = images_ambiguity  # imgs[np.random.randint(0,len(imgs))]

            GT_path = "{}/{}/{}/{}".format(self.root_path, 'Images', record['subfolder'], GT_path)
            image_feature_ambiguity = self.pickle_dict[GT_path]
            content_feature_ambiguity = self.pickle_dict[content]

            content_feature_ambiguity = torch.from_numpy(content_feature_ambiguity).float()
            image_feature_ambiguity = torch.from_numpy(image_feature_ambiguity).float()

        if not self.with_ambiguity:
            return (content_feature, image_feature, label, 0), (GT_path)
        else:
            return (content_feature, image_feature, label, 0), (GT_path), \
                   (content_feature_ambiguity, image_feature_ambiguity, label_ambiguity)

    def __len__(self):
        return len(self.label_dict)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t

if __name__ == '__main__':
    filename = "../train_gossip.pickle"
    with open(filename, 'rb') as f:
        pickle_dict = pickle.load(f)
    print(list(pickle_dict.keys())[7000:7010])