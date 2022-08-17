
import cv2

import torch
import torch.utils.data as data
import data.util as util

import torchvision.transforms.functional as F
from torchvision import datasets, transforms
# datasets.ImageFolder

from PIL import Image
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from tqdm import tqdm
from transformers import BertModel, BertTokenizer
#import clip

# __all__ = [
#     "Compose",
#     "ToTensor",
#     "PILToTensor",
#     "ConvertImageDtype",
#     "ToPILImage",
#     "Normalize",
#     "Resize",
#     "CenterCrop",
#     "Pad",
#     "Lambda",
#     "RandomApply",
#     "RandomChoice",
#     "RandomOrder",
#     "RandomCrop",
#     "RandomHorizontalFlip",
#     "RandomVerticalFlip",
#     "RandomResizedCrop",
#     "FiveCrop",
#     "TenCrop",
#     "LinearTransformation",
#     "ColorJitter",
#     "RandomRotation",
#     "RandomAffine",
#     "Grayscale",
#     "RandomGrayscale",
#     "RandomPerspective",
#     "RandomErasing",
#     "GaussianBlur",
#     "InterpolationMode",
#     "RandomInvert",
#     "RandomPosterize",
#     "RandomSolarize",
#     "RandomAdjustSharpness",
#     "RandomAutocontrast",
#     "RandomEqualize",
# ]
import random

class weibo_dataset_test(data.Dataset):

    def __init__(self, root_path='/home/groupshare/bootstrap_FND_test/weibo_modification_test_liu/',
                 excel_path='test1_weibo.xlsx',
                 image_size=224):
        super(weibo_dataset_test, self).__init__()
        self.root_path = root_path
        self.index = 0
        self.label_dict = []
        self.image_size = image_size

        wb = openpyxl.load_workbook(self.root_path+excel_path)
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row


        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet['B' + str(i)].value)
            label = 0
            content = str(sheet['C' + str(i)].value)
            # imgs = images_name.split('|')
            record = {}
            record['images'] = images_name
            # print(images_name)
            record['label'] = label
            record['content'] = content
            # new 9:1 division
            # if (is_train and i%10!=9) or (not is_train and i%10==9):
            self.label_dict.append(record)

        assert len(self.label_dict)!=0, 'Error: GT path is empty.'

    def __getitem__(self, index):

        GT_size = self.image_size

        # get GT image
        record = self.label_dict[index]
        GT_path, label, content = record['images'], record['label'], record['content']

        try:

            GT_path = "{}/{}".format(self.root_path,GT_path)
            img_GT = util.read_img(GT_path)
        except Exception:
            raise IOError("Load {} Error {}".format(GT_path, record['images']))

        img_GT = util.channel_convert(img_GT.shape[2], 'RGB', [img_GT])[0]

        ###### directly resize instead of crop
        img_GT = cv2.resize(np.copy(img_GT), (GT_size, GT_size),
                            interpolation=cv2.INTER_LINEAR)

        orig_height, orig_width, _ = img_GT.shape
        H, W, _ = img_GT.shape

        # BGR to RGB, HWC to CHW, numpy to tensor
        if img_GT.shape[2] == 3:
            img_GT = img_GT[:, :, [2, 1, 0]]


        img_GT = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))).float()


        return (content, img_GT, label, 0), (GT_path)

        # 如果用smooth label来训练网络， 则上面的label应该改成 torch.tensor(label,dtype=torch.float32)

    def __len__(self):
        return len(self.label_dict)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t
