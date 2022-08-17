
import cv2

import torch
import torch.utils.data as data
import data.util as util
import random
import torchvision.transforms.functional as F

from PIL import Image
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from tqdm import tqdm
from transformers import BertModel, BertTokenizer
import clip

class Twitter_dataset(data.Dataset):

    def __init__(self, root_path='/home/groupshare/Twitter', image_size=224, is_train=True, remove_duplicate=True,
                 is_sample_positive=0.5):
        super(Twitter_dataset, self).__init__()
        self.remove_duplicate = remove_duplicate
        self.is_train = is_train
        self.root_path = root_path
        self.is_sample_positive = is_sample_positive
        self.index = 0
        # self.text_max_len = 170
        # self.token = BertTokenizer.from_pretrained('bert-base-uncased')
        self.label_dict = []
        self.image_size = image_size
        self.image_path = self.root_path+'/images/images' if is_train else self.root_path+'/image-verification-corpus-master/mediaeval2016/testset/Mediaeval2016_TestSet_Images/Mediaeval2016_TestSet_Images'
        print("{} path: {}".format("Train" if is_train else "Test",self.image_path))
        list_dirs = os.listdir(self.image_path)
        self.image_listdir = {}
        for list_dir in list_dirs:
            idx = list_dir.rfind('.')
            self.image_listdir[list_dir[:idx]] = list_dir

        train_path = root_path + '/train_twitters.xlsx'
        test_path = root_path+'/test_datasets.xlsx'
        wb = openpyxl.load_workbook(train_path if is_train else test_path)
        self.IMAGE_DICT, self.IMAGE_DICT_KEYS = {}, []
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet['E' + str(i)].value)
            label = int(sheet['H' + str(i)].value)
            # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
            label = 1 if label==0 else 0

            content = str(sheet['C' + str(i)].value)
            # imgs = images_name.split('|')
            record = {}
            record['images'] = images_name
            # print(images_name)
            record['label'] = label
            record['content'] = []
            record['content'].append(content)
            # PREVIOUSLY, WE JUST APPEND EVERY RECORDS AS FOLLOWS
            if not self.is_train or not self.remove_duplicate:
                self.label_dict.append(record)
            # NOW, WE ACCUMULATE NEWS USING SAME IMAGES IN INDIVIDUAL CLUSTERS
            # AND EACH EPOCH WILL ONLY WALK THROUGH A SINGLE NEWS WITHIN EACH CLUSTER
            else:
                key = images_name+"_"+str(label)
                if key not in self.IMAGE_DICT:
                    self.IMAGE_DICT[key] = record
                    self.label_dict.append(record)
                else:
                    if self.is_sample_positive < 1 and label == 1 and random.random() > self.is_sample_positive:
                        continue
                    record_previous = self.IMAGE_DICT[key]
                    record_previous['content'].append(content)

        # if not self.is_train or not self.remove_duplicate:
        assert len(self.label_dict)!=0, 'Error: GT path is empty.'
        # else:
        #     assert len(self.IMAGE_DICT_KEYS)!=0, 'Error: GT path is empty.'

        # if use_extra_data:
        #     wb = openpyxl.load_workbook('/home/groupshare/gossip/gossip_train_no_filt.xlsx')
        #     sheetnames = wb.sheetnames
        #     sheet = wb[sheetnames[0]]
        #     rows = sheet.max_row
        #     for i in tqdm(range(2, rows + 1)):
        #         images_name = str(sheet['C' + str(i)].value)
        #         label = int(sheet['D' + str(i)].value)
        #         # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
        #         label = 1 if label == 0 else 0
        #         if label == 0 and random.random() > 0.4:
        #             continue
        #         content = str(sheet['B' + str(i)].value)
        #         category = str(sheet['E' + str(i)].value)
        #         ##  NEWS TYPE: 0 MULTIMODAL 1 IMAGE 2 TEXT
        #         if not is_use_unimodal or "multi" in category:
        #             category = 0
        #         elif "image" in category:
        #             category = 1
        #         else:
        #             category = 2
        #         # imgs = images_name.split('|')
        #         record = {}
        #         record['images'] = images_name
        #         record['label'] = label
        #         record['content'] = content
        #         record['subfolder'] = '{}_{}'.format(self.dataset_name, 'train' if is_train else 'test')
        #         record['category'] = category
        #         self.label_dict.append(record)

    def __getitem__(self, index):

        GT_size = self.image_size # 这个自己指定一下

        # get GT image
        # if not self.is_train or not self.remove_duplicate:
        record = self.label_dict[index]
        # else:
        #     key = self.IMAGE_DICT_KEYS[index]
        #     queried_list = self.IMAGE_DICT[key]
        #     idx_rand = random.randint(0,len(queried_list)-1)
        #     record = queried_list[idx_rand]
        images, label, content_list = record['images'], record['label'], record['content']

        idx_rand = random.randint(0, len(content_list) - 1)
        content = content_list[idx_rand]

        # imgs = images.split('|')
        imgs = images.split(',')
        GT_path = imgs[np.random.randint(0, len(imgs))]
        # if '/' in GT_path:
        #     # excel中文件名前面可能会有'(non)rumor_images/'，这个在读图的时候是不需要的，直接过滤
        #     GT_path = GT_path[GT_path.rfind('/')+1:]
        try:
            is_found_image = False
            if GT_path in self.image_listdir:
                GT_path = self.image_listdir[GT_path]
                is_found_image = True
            # Sandy extra care
            if not is_found_image and GT_path[:5] == 'sandy':
                image_item_A, image_item_B = 'sandyA' + GT_path[5:], 'sandyB' + GT_path[5:]
                if image_item_A in self.image_listdir:
                    GT_path = self.image_listdir[image_item_A]
                elif image_item_B in self.image_listdir:
                    GT_path = self.image_listdir[image_item_B]

            GT_path = "{}/{}".format(self.image_path,GT_path)
            img_GT = util.read_img(GT_path)
        except Exception:
            raise IOError("Load {} Error {}".format(GT_path, record['images']))

        img_GT = util.channel_convert(img_GT.shape[2], 'RGB', [img_GT])[0]

        ###### directly resize instead of crop
        img_GT = cv2.resize(np.copy(img_GT), (GT_size, GT_size),
                            interpolation=cv2.INTER_LINEAR)

        orig_height, orig_width, _ = img_GT.shape
        H, W, _ = img_GT.shape

        # BGR to RGB, HWC to CHW, numpy to tensor
        if img_GT.shape[2] == 3:
            img_GT = img_GT[:, :, [2, 1, 0]]


        img_GT = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))).float()
        ### Error: stack expects each tensor to be equal size, but got [11, 170] at entry 0 and [161, 170] at entry 1
        # data = self.token.batch_encode_plus(batch_text_or_text_pairs=content,
        #                                     truncation=True,
        #                                     padding='max_length',
        #                                     max_length=self.text_max_len,
        #                                     return_tensors='pt',
        #                                     return_length=True)
        #
        # input_ids = data['input_ids']
        # attention_mask = data['attention_mask']
        # token_type_ids = data['token_type_ids']

        return (content, img_GT), label
        # 如果用smooth label来训练网络， 则上面的label应该改成 torch.tensor(label,dtype=torch.float32)

    def __len__(self):
        # if not self.is_train or not self.remove_duplicate:
        return len(self.label_dict)
        # NOW IN EACH EPOCH WE GO THROUGH CLUSTERS AND SAMPLE ONLY ONE NEWS
        # else:
        #     return len(self.IMAGE_DICT_KEYS)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t


    # def collate_fn(self,data):
    #     sents = [i[0][0] for i in data]
    #     image = [i[0][1] for i in data]
    #     imageclip = [i[0][2] for i in data]
    #     textclip = [i[0][3] for i in data]
    #     labels = [i[1] for i in data]
    #
    #     # 编码
    #
    #     image = torch.stack(image)
    #
    #     labels = torch.LongTensor(labels)
    #
    #     # print(data['length'], data['length'].max())
    #
    #     return input_ids, attention_mask, token_type_ids, image, imageclip, textclip, labels

