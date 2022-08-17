
import cv2

import torch
import torch.utils.data as data
import data.util as util

import torchvision.transforms.functional as F

from PIL import Image
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from tqdm import tqdm
from transformers import BertModel, BertTokenizer
import clip

class MixSet_dataset(data.Dataset):

    def __init__(self, root_path='/home/groupshare/MixSet', image_size=512, is_train=True):
        super(MixSet_dataset, self).__init__()
        self.index = 0
        self.text_max_len = 170
        # self.token = BertTokenizer.from_pretrained('bert-base-chinese')
        self.label_dict = []
        self.image_size = image_size
        wb = openpyxl.load_workbook(root_path+'/cleaned_data.xlsx')
        self.nonrumor_folder = root_path+'/nonrumor_images/'
        self.rumor_folder = root_path+'/rumor_images/'
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet['F' + str(i)].value)
            label = int(sheet['C' + str(i)].value)
            # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
            label = 0 if label == 1 else 1
            content = str(sheet['E' + str(i)].value)
            # imgs = images_name.split('|')
            record = {}
            record['images'] = images_name
            record['label'] = label
            record['content'] = content
            # new 9:1 division
            if (is_train and i%10!=9) or (not is_train and i%10==9):
                self.label_dict.append(record)

        # 9:1 division
        # division_idx = int(0.9 * len(self.label_dict))
        # if is_train:
        #     self.label_dict = self.label_dict[:division_idx]
        # else:
        #     self.label_dict = self.label_dict[division_idx:]
        assert len(self.label_dict)!=0, 'Error: GT path is empty.'

    def __getitem__(self, index):

        GT_size = self.image_size # 这个自己指定一下

        # get GT image
        record = self.label_dict[index]
        images, label, content = record['images'], record['label'], record['content']
        imgs = images.split('|')
        GT_path = imgs[np.random.randint(0,len(imgs))]
        if '/' in GT_path:
            # excel中文件名前面可能会有'(non)rumor_images/'，这个在读图的时候是不需要的，直接过滤
            GT_path = GT_path[GT_path.rfind('/')+1:]
        try:
            GT_path = "{}{}".format(self.nonrumor_folder if label==0 else self.rumor_folder, GT_path)
            img_GT = util.read_img(GT_path)
        except Exception:
            raise IOError("Load {} Error".format(GT_path))

        img_GT = util.channel_convert(img_GT.shape[2], 'RGB', [img_GT])[0]

        ###### directly resize instead of crop
        img_GT = cv2.resize(np.copy(img_GT), (GT_size, GT_size),
                            interpolation=cv2.INTER_LINEAR)

        orig_height, orig_width, _ = img_GT.shape
        H, W, _ = img_GT.shape

        # BGR to RGB, HWC to CHW, numpy to tensor
        if img_GT.shape[2] == 3:
            img_GT = img_GT[:, :, [2, 1, 0]]


        img_GT = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))).float()
        ### Error: stack expects each tensor to be equal size, but got [11, 170] at entry 0 and [161, 170] at entry 1
        # data = self.token.batch_encode_plus(batch_text_or_text_pairs=content,
        #                                     truncation=True,
        #                                     padding='max_length',
        #                                     max_length=self.text_max_len,
        #                                     return_tensors='pt',
        #                                     return_length=True)
        #
        # input_ids = data['input_ids']
        # attention_mask = data['attention_mask']
        # token_type_ids = data['token_type_ids']

        return (content, img_GT), label
        # 如果用smooth label来训练网络， 则上面的label应该改成 torch.tensor(label,dtype=torch.float32)

    def __len__(self):
        return len(self.label_dict)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t


    # def collate_fn(self,data):
    #     sents = [i[0][0] for i in data]
    #     image = [i[0][1] for i in data]
    #     imageclip = [i[0][2] for i in data]
    #     textclip = [i[0][3] for i in data]
    #     labels = [i[1] for i in data]
    #
    #     # 编码
    #
    #     image = torch.stack(image)
    #
    #     labels = torch.LongTensor(labels)
    #
    #     # print(data['length'], data['length'].max())
    #
    #     return input_ids, attention_mask, token_type_ids, image, imageclip, textclip, labels

