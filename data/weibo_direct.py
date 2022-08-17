import pickle
import cv2

import torch
import torch.utils.data as data
import data.util as util

import torchvision.transforms.functional as F
from torchvision import datasets, transforms
# datasets.ImageFolder

from PIL import Image
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from tqdm import tqdm
from transformers import BertModel, BertTokenizer
import clip

# __all__ = [
#     "Compose",
#     "ToTensor",
#     "PILToTensor",
#     "ConvertImageDtype",
#     "ToPILImage",
#     "Normalize",
#     "Resize",
#     "CenterCrop",
#     "Pad",
#     "Lambda",
#     "RandomApply",
#     "RandomChoice",
#     "RandomOrder",
#     "RandomCrop",
#     "RandomHorizontalFlip",
#     "RandomVerticalFlip",
#     "RandomResizedCrop",
#     "FiveCrop",
#     "TenCrop",
#     "LinearTransformation",
#     "ColorJitter",
#     "RandomRotation",
#     "RandomAffine",
#     "Grayscale",
#     "RandomGrayscale",
#     "RandomPerspective",
#     "RandomErasing",
#     "GaussianBlur",
#     "InterpolationMode",
#     "RandomInvert",
#     "RandomPosterize",
#     "RandomSolarize",
#     "RandomAdjustSharpness",
#     "RandomAutocontrast",
#     "RandomEqualize",
# ]
import random

class weibo_dataset(data.Dataset):

    def __init__(self, root_path='/home/groupshare/weibo', image_size=224, is_train=True,
                 with_ambiguity=False,use_soft_label=False, is_use_unimodal=False):
        super(weibo_dataset, self).__init__()
        self.with_ambiguity = with_ambiguity
        self.use_soft_label = use_soft_label
        self.label_ambiguity = []
        self.is_train = is_train
        print("Using AMBIGUITY LEARNING: {}".format(self.with_ambiguity))
        self.root_path = root_path
        self.index = 0
        self.label_dict = []
        self.image_size = image_size
        self.transform = transforms.Compose([
        # transforms.Resize(input_size),
        # transforms.CenterCrop(input_size),
        # transforms.RandomAffine(degrees=0, translate=(0.05, 0.05)),
            transforms.RandomVerticalFlip(),
            transforms.RandomHorizontalFlip(),
        # transforms.ToTensor(),
        # transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),

         ])

        self.pickle_dict = None
        filename = f'{"train" if is_train else "val"}_weibo.pickle'
        with open(filename,'rb') as f:
            self.pickle_dict = pickle.load(f)

        wb = openpyxl.load_workbook(root_path+'/{}_datasets_WWW.xlsx'.format('train' if is_train else 'test'))
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet['F' + str(i)].value)
            label = int(sheet['C' + str(i)].value)
            # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
            label = 1 if label==0 else 0
            content = str(sheet['E' + str(i)].value)
            # imgs = images_name.split('|')
            record = {}
            record['images'] = images_name
            # print(images_name)
            record['label'] = label
            record['content'] = content
            # new 9:1 division
            # if (is_train and i%10!=9) or (not is_train and i%10==9):
            self.label_dict.append(record)

        # AMBIGUITY LEARNING
        if self.is_train and self.with_ambiguity:
            wb = openpyxl.load_workbook(root_path + '/{}.xlsx'
                                        .format('weibo_train_ambiguity'))
            sheetnames = wb.sheetnames
            sheet = wb[sheetnames[0]]
            rows = sheet.max_row
            for i in tqdm(range(2, rows + 1)):
                images_name = str(sheet['C' + str(i)].value)
                label = int(sheet['D' + str(i)].value)
                # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
                # label = 1 if label == 0 else 0
                content = str(sheet['B' + str(i)].value)
                category = str(sheet['E' + str(i)].value)
                ##  NEWS TYPE: 0 MULTIMODAL 1 IMAGE 2 TEXT
                if not is_use_unimodal or "multi" in category:
                    category = 0
                elif "image" in category:
                    category = 1
                else:
                    category = 2
                # imgs = images_name.split('|')
                record = {}
                record['images'] = images_name
                record['label'] = label
                record['content'] = content
                # record['subfolder'] = '{}_{}'.format(self.dataset_name, 'train' if is_train else 'test')
                record['category'] = category
                self.label_ambiguity.append(record)

        assert len(self.label_dict)!=0, 'Error: GT path is empty.'

    def __getitem__(self, index):

        GT_size = self.image_size # 这个自己指定一下


        record = self.label_dict[index]
        images, label, content = record['images'], record['label'], record['content']
        imgs = images.split('|')
        while True:
            GT_path = imgs[np.random.randint(0, len(imgs))]
            if GT_path in self.pickle_dict:
                image_feature = self.pickle_dict[GT_path]
                break
        content_feature = self.pickle_dict[content]

        content_feature = torch.from_numpy(content_feature).float()
        image_feature = torch.from_numpy(image_feature).float()

        # AMBIGUITY: SAME, REPEATED CODE
        if self.with_ambiguity:
            # get GT image
            index1 = random.randint(0, len(self.label_ambiguity) - 1)
            record = self.label_ambiguity[index1]
            images_ambiguity, label_ambiguity, content_ambiguity = record['images'], record['label'], record['content']
            if self.use_soft_label:
                if label_ambiguity == 0:
                    label_ambiguity = 0.2
                else:
                    label_ambiguity = 0.8

            imgs = images_ambiguity.split('|')
            while True:
                GT_path = imgs[np.random.randint(0, len(imgs))]
                if GT_path in self.pickle_dict:
                    image_feature_ambiguity = self.pickle_dict[GT_path]
                    break
            content_feature_ambiguity = self.pickle_dict[content]

            content_feature_ambiguity = torch.from_numpy(content_feature_ambiguity).float()
            image_feature_ambiguity = torch.from_numpy(image_feature_ambiguity).float()

        if not self.with_ambiguity:
            return (content_feature, image_feature, label, 0), (GT_path)
        else:
            return (content_feature, image_feature, label, 0), (GT_path), \
                   (content_feature_ambiguity, image_feature_ambiguity, label_ambiguity)

        # 如果用smooth label来训练网络， 则上面的label应该改成 torch.tensor(label,dtype=torch.float32)

    def __len__(self):
        return len(self.label_dict)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t


    # def collate_fn(self,data):
    #     sents = [i[0][0] for i in data]
    #     image = [i[0][1] for i in data]
    #     imageclip = [i[0][2] for i in data]
    #     textclip = [i[0][3] for i in data]
    #     labels = [i[1] for i in data]
    #
    #     # 编码
    #
    #     image = torch.stack(image)
    #
    #     labels = torch.LongTensor(labels)
    #
    #     # print(data['length'], data['length'].max())
    #
    #     return input_ids, attention_mask, token_type_ids, image, imageclip, textclip, labels

