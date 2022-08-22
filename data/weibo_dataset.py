import copy

import cv2

import torch
import torch.utils.data as data
import data.util as util

import torchvision.transforms.functional as F
from torchvision import datasets, transforms
# datasets.ImageFolder
import albumentations as A
from PIL import Image
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from tqdm import tqdm
from transformers import BertModel, BertTokenizer
#import clip

# __all__ = [
#     "Compose",
#     "ToTensor",
#     "PILToTensor",
#     "ConvertImageDtype",
#     "ToPILImage",
#     "Normalize",
#     "Resize",
#     "CenterCrop",
#     "Pad",
#     "Lambda",
#     "RandomApply",
#     "RandomChoice",
#     "RandomOrder",
#     "RandomCrop",
#     "RandomHorizontalFlip",
#     "RandomVerticalFlip",
#     "RandomResizedCrop",
#     "FiveCrop",
#     "TenCrop",
#     "LinearTransformation",
#     "ColorJitter",
#     "RandomRotation",
#     "RandomAffine",
#     "Grayscale",
#     "RandomGrayscale",
#     "RandomPerspective",
#     "RandomErasing",
#     "GaussianBlur",
#     "InterpolationMode",
#     "RandomInvert",
#     "RandomPosterize",
#     "RandomSolarize",
#     "RandomAdjustSharpness",
#     "RandomAutocontrast",
#     "RandomEqualize",
# ]
import random

class weibo_dataset(data.Dataset):

    def __init__(self,
                 dataset='weibo', image_size=224, is_train=True,
                 with_ambiguity=False,use_soft_label=False, is_use_unimodal=False,
                 not_on_12=0
                 ):
        super(weibo_dataset, self).__init__()
        self.not_valid_set = set()
        not_on_12 = not_on_12>0
        print(f"not on 12? {not_on_12}")
        self.with_ambiguity = with_ambiguity
        self.use_soft_label = use_soft_label
        self.label_ambiguity = []
        self.is_train = is_train
        print("Using AMBIGUITY LEARNING: {}".format(self.with_ambiguity))
        self.dataset = dataset
        root_path = '/home/groupshare/'+self.dataset
        self.root_path = root_path if not not_on_12 else root_path[5:]
        if '21' in self.root_path:
            print("We are using Weibo 21.")
        self.root_path_ambiguity = root_path #'/home/groupshare/weibo' if '21' not in self.root_path else '/home/groupshare/Weibo_21'
        self.root_path_ambiguity = root_path if not not_on_12 else root_path[5:]
        self.ambiguity_excel = f'{self.root_path_ambiguity}/{self.dataset}_train_ambiguity_new.xlsx'
        self.index = 0
        self.label_dict = []
        self.image_size = image_size
        self.transform_just_resize = A.Compose(
            [
                A.Resize(always_apply=True, height=image_size, width=image_size)
            ]
        )
        self.transform = A.Compose(
            [
                A.HorizontalFlip(p=0.5),
                # A.ElasticTransform(p=0.5),
                A.OneOf(
                    [
                        A.CLAHE(always_apply=False, p=0.25),
                        A.RandomBrightnessContrast(always_apply=False, p=0.25),
                        A.Equalize(always_apply=False, p=0.25),
                        A.RGBShift(always_apply=False, p=0.25),
                    ]
                ),
                A.OneOf(
                    [
                        A.ImageCompression(always_apply=False, quality_lower=60, quality_upper=100, p=0.2),
                        A.MedianBlur(always_apply=False, p=0.2),
                        A.GaussianBlur(always_apply=False, p=0.2),
                        # A.MotionBlur(always_apply=False, p=0.2),
                        A.GaussNoise(always_apply=False, p=0.2),
                        A.ISONoise(always_apply=False, p=0.2)
                    ]
                ),
                A.Resize(always_apply=True,height=image_size, width=image_size)
            ]
        )

        wb = openpyxl.load_workbook(f"{self.root_path}/{'train' if self.is_train else 'test'}_datasets{'' if '21' in self.root_path else '_WWW_new'}.xlsx")

        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row

        fake_news_num = 0
        ## calculate true:fake
        for i in tqdm(range(2, rows + 1)):
            label = int(sheet['C' + str(i)].value)
            # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
            label = 1 if label == 0 else 0
            fake_news_num += label

        downsample_rate = fake_news_num / (rows - fake_news_num)
        print(f"Downsample rate: {downsample_rate}")


        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet['F' + str(i)].value)
            label = int(sheet['C' + str(i)].value)
            # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
            label = 1 if label==0 else 0
            content = str(sheet['E' + str(i)].value)
            # imgs = images_name.split('|')
            record = {}
            record['images'] = images_name
            # print(images_name)
            record['label'] = label
            record['content'] = content
            # new 9:1 division
            # if (is_train and i%10!=9) or (not is_train and i%10==9):
            self.label_dict.append(record)

        # AMBIGUITY LEARNING
        if self.is_train and self.with_ambiguity:

            wb = openpyxl.load_workbook(self.ambiguity_excel)
            sheetnames = wb.sheetnames
            sheet = wb[sheetnames[0]]
            rows = sheet.max_row
            for i in tqdm(range(2, rows + 1)):
                images_name = str(sheet['C' + str(i)].value)
                label = int(sheet['D' + str(i)].value)
                # 注意： 收集MixSet的时候，把真新闻标记为1，因此最好把它反过来
                # label = 1 if label == 0 else 0
                content = str(sheet['B' + str(i)].value)
                category = str(sheet['E' + str(i)].value)
                ##  NEWS TYPE: 0 MULTIMODAL 1 IMAGE 2 TEXT
                if not is_use_unimodal or "multi" in category:
                    category = 0
                elif "image" in category:
                    category = 1
                else:
                    category = 2
                # imgs = images_name.split('|')
                record = {}
                record['images'] = images_name
                record['label'] = label
                record['content'] = content
                # record['subfolder'] = '{}_{}'.format(self.dataset_name, 'train' if is_train else 'test')
                record['category'] = category
                self.label_ambiguity.append(record)

        assert len(self.label_dict)!=0, 'Error: GT path is empty.'

    def __getitem__(self, index):

        GT_size = self.image_size # 这个自己指定一下
        find_path, img_GT = False, None
        while not find_path:
            # get GT image
            record = self.label_dict[index]
            images, label, content = record['images'], record['label'], record['content']
            imgs = images.split('|')
            for index_image in random.sample(range(len(imgs)),len(imgs)):
                GT_path = imgs[index_image]

                GT_path = "{}/{}".format(self.root_path, GT_path)
                find_path = os.path.exists(GT_path)
                if not find_path:
                    print(f"File not found!{GT_path}")
                elif not GT_path in self.not_valid_set:
                    img_GT = cv2.imread(GT_path, cv2.IMREAD_COLOR)
                    if img_GT is None:
                        print(f"File cannot open!{GT_path}")
                        find_path = False
                    else:
                        H_origin, W_origin, _ = img_GT.shape
                        if H_origin < 100 or W_origin < 100 or H_origin / W_origin < 0.33 or H_origin / W_origin > 3:  # 'text' in category:
                            # print(f"Unimodal text detected {H_origin} {W_origin}. Skip.")
                            find_path = False
                            # self.not_valid_set.add(GT_path)
                            # img_GT = torch.zeros_like(img_GT)
                        elif len(content)<10:
                            print("Find length not satisfying")
                            # content = "No image provided for this news"
                            find_path = False
                            self.not_valid_set.add(GT_path)
                            break
                        else:
                            find_path = True
                            if img_GT.ndim == 2:
                                img_GT = np.expand_dims(img_GT, axis=2)
                            # some images have 4 channels
                            if img_GT.shape[2] > 3:
                                img_GT = img_GT[:, :, :3]

                            img_GT = util.channel_convert(img_GT.shape[2], 'RGB', [img_GT])[0]
                            break
            index = np.random.randint(0, len(self.label_dict))

        if not self.is_train:
            img_GT_augment = self.transform_just_resize(image=copy.deepcopy(img_GT))["image"]
        else:
            img_GT_augment = self.transform(image=copy.deepcopy(img_GT))["image"]
        img_GT = self.transform_just_resize(image=copy.deepcopy(img_GT))["image"]
        img_GT = img_GT.astype(np.float32) / 255.
        img_GT_augment = img_GT_augment.astype(np.float32) / 255.

        ###### directly resize instead of crop
        # img_GT = cv2.resize(np.copy(img_GT), (GT_size, GT_size),
        #                     interpolation=cv2.INTER_LINEAR)

        orig_height, orig_width, _ = img_GT.shape
        H, W, _ = img_GT.shape

        # BGR to RGB, HWC to CHW, numpy to tensor
        if img_GT.shape[2] == 3:
            img_GT = img_GT[:, :, [2, 1, 0]]
        if img_GT_augment.shape[2] == 3:
            img_GT_augment = img_GT_augment[:, :, [2, 1, 0]]

        img_GT = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))).float()
        img_GT_augment = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT_augment, (2, 0, 1)))).float()

        # AMBIGUITY: SAME, REPEATED CODE
        if self.with_ambiguity:
            # get GT image
            index1 = random.randint(0, len(self.label_ambiguity) - 1)
            record = self.label_ambiguity[index1]
            images_ambiguity, label_ambiguity, content_ambiguity = record['images'], record['label'], record['content']
            if self.use_soft_label:
                if label_ambiguity == 0:
                    label_ambiguity = 0.2
                else:
                    label_ambiguity = 0.8

            imgs = images_ambiguity.split('|')
            GT_path = imgs[np.random.randint(0, len(imgs))]
            # if '/' in GT_path:
            #     # excel中文件名前面可能会有'(non)rumor_images/'，这个在读图的时候是不需要的，直接过滤
            #     GT_path = GT_path[GT_path.rfind('/')+1:]
            try:

                GT_path = "{}/{}".format(self.root_path_ambiguity, GT_path)
                img_ambiguity = util.read_img(GT_path)


                img_ambiguity = util.channel_convert(img_ambiguity.shape[2], 'RGB', [img_ambiguity])[0]

                ###### directly resize instead of crop
                img_ambiguity = cv2.resize(np.copy(img_ambiguity), (GT_size, GT_size),
                                           interpolation=cv2.INTER_LINEAR)

                orig_height, orig_width, _ = img_ambiguity.shape
                H, W, _ = img_ambiguity.shape

                # BGR to RGB, HWC to CHW, numpy to tensor
                if img_ambiguity.shape[2] == 3:
                    img_ambiguity = img_ambiguity[:, :, [2, 1, 0]]

                img_ambiguity = torch.from_numpy(
                    np.ascontiguousarray(np.transpose(img_ambiguity, (2, 0, 1)))).float()
            except Exception:
                print(f"[Exception] load image error at {GT_path}. Using a zero-matrix instead")
                img_ambiguity = torch.zeros((3, GT_size, GT_size))

        if not self.with_ambiguity:
            return (content, img_GT, img_GT_augment, label, 0), (GT_path)
        else:
            return (content, img_GT, img_GT_augment, label, 0), (GT_path), (content_ambiguity, img_ambiguity, label_ambiguity)

        # 如果用smooth label来训练网络， 则上面的label应该改成 torch.tensor(label,dtype=torch.float32)

    def __len__(self):
        return len(self.label_dict)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t


    # def collate_fn(self,data):
    #     sents = [i[0][0] for i in data]
    #     image = [i[0][1] for i in data]
    #     imageclip = [i[0][2] for i in data]
    #     textclip = [i[0][3] for i in data]
    #     labels = [i[1] for i in data]
    #
    #     # 编码
    #
    #     image = torch.stack(image)
    #
    #     labels = torch.LongTensor(labels)
    #
    #     # print(data['length'], data['length'].max())
    #
    #     return input_ids, attention_mask, token_type_ids, image, imageclip, textclip, labels

