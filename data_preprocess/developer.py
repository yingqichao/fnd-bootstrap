import time
import json
from tkinter import *
from PIL import Image, ImageTk
import wget
import copy
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from scp import SCPClient
from tqdm import tqdm
import datetime
import random
from tkinter.filedialog import askdirectory, askopenfilename
import shutil
import options.options as option
import photohash
import cv2
import pickle

# from imagededup.methods import PHash
# phasher = PHash()

from variables import rumor_root, weibo_fake_root, weibo_real_root, weibo21_fake_root, weibo21_real_root, \
    get_workbook


def get_image_list(dataset_name,do_download=False):
    native_image_list_dict, native_image_list = {}, []
    news_xlsx = './dataset/{}/news_datasets.xlsx'.format(dataset_name)
    all_images_xlsx = './dataset/{}/all_images.xlsx'.format(dataset_name)
    # rumor_dataset
    if dataset_name == "rumor_dataset":
        for l_num in range(2):
            all_images_xlsx = './dataset/{}/all_images_{}.xlsx'.format(dataset_name,"chinese" if l_num==0 else "english")
            native_image_list_dict, native_image_list = {}, []
            folder = "E:/rumor_dataset/rumor_datasets/images/All images/"
            images = os.listdir(folder)
            images_set = set(images)
            wb = openpyxl.load_workbook(news_xlsx)
            sheetnames = wb.sheetnames
            sheet = wb[sheetnames[0]]
            rows = sheet.max_row
            for i in tqdm(range(2, rows + 1)):
                if (l_num==0 and ((i>=1002 and i<=5001) or (i>=6002))) or (l_num==1 and ((i<1001) or (i>5001 and i<6002))):
                    # Chinese/English
                    image_names = sheet['E' + str(i)].value
                    if len(image_names)!=0:
                        num_addded, image_name_list = 0, set()
                        image_name_list_1 = image_names.split('\t')
                        image_name_list_2 = image_names.split(';')
                        if len(image_name_list_1)>=len(image_name_list_2):
                            image_name_list = set(image_name_list_1)
                        else:
                            image_name_list = set(image_name_list_2)
                        for image_name in image_name_list:
                            image_name = image_name.lstrip(' ').rstrip(' ')
                            native_image_list_dict = {}
                            if image_name in images_set:
                                native_image_list_dict['path'] = copy.copy(image_name)
                                native_image_list_dict['label'] = int(sheet['F' + str(i)].value)
                                native_image_list_dict['index'] = i
                                native_image_list.append(native_image_list_dict)
                                num_addded += 1
                        if len(image_name_list)>=1:
                            print("{}: {}/{} {}".format(i,num_addded,len(image_name_list),image_names))
            df = pd.DataFrame(native_image_list)
            df.to_excel(all_images_xlsx)
    else:
        all_images_xlsx = './dataset/{}/all_images.xlsx'.format(dataset_name)
        native_image_list_dict, native_image_list = {}, []
        folder1, folder2 = "E:/{}/nonrumor_images".format(dataset_name), "E:/{}/rumor_images".format(dataset_name)
        images1, images2 = set(os.listdir(folder1)), set(os.listdir(folder2))
        images_set = images1.union(images2)
        wb = openpyxl.load_workbook(news_xlsx)
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            image_names = sheet['C' + str(i)].value
            label = int(sheet['E' + str(i)].value)
            if image_names is not None and len(image_names) != 0:
                num_addded, image_name_list = 0, set()
                image_name_list = image_names.split('|')
                for image_name in image_name_list:
                    full_image_name = image_name.lstrip(' ').rstrip(' ') # http url
                    short_name = full_image_name[full_image_name.rfind('/')+1:]
                    native_image_list_dict = {}
                    if do_download and short_name not in images_set and full_image_name[:4]=="http":
                        full_name = 'E:/{}/'.format(dataset_name) + ("rumor_images/" if label == 0 else "nonrumor_images/") + short_name
                        try:
                            wget.download(full_image_name, 'E:/weibo/Downloaded/'+short_name)
                            shutil.copy('E:/{}/Downloaded/'.format(dataset_name)+short_name, full_name)
                            images_set.add(short_name)
                            print("Download ok. {}".format(full_name))
                        except Exception:
                            print("Download error. {}".format(full_name))

                    if short_name in images_set:

                        native_image_list_dict['path'] = ("rumor_images/" if label==0 else "nonrumor_images/")+short_name
                        native_image_list_dict['label'] = label
                        native_image_list_dict['index'] = i
                        native_image_list.append(native_image_list_dict)
                        num_addded += 1
                if len(image_name_list) >= 1:
                    print("{}: {}/{} {}".format(i, num_addded, len(image_name_list), image_names))

        df = pd.DataFrame(native_image_list)
        df.to_excel(all_images_xlsx)




def check_similarity(target_folder,target_xlsx, dataset_name):
    images1, images2 = {}, {}
    # wb = openpyxl.load_workbook(xlsx1)
    # sheetnames = wb.sheetnames
    # sheet = wb[sheetnames[0]]
    # rows = sheet.max_row
    # for i in tqdm(range(2, rows + 1)):
    #     image_name = sheet['B' + str(i)].value
    #     hash = photohash.average_hash(folder1+image_name)
    #     images1[folder1+image_name] = hash

    # from imagededup.methods import CNN
    # cnn = CNN()
    # encodings = cnn.encode_images(image_dir=folder1)
    # print("Start finding dups...")
    # duplicates = cnn.find_duplicates(encoding_map=encodings)
    # print("End finding dups...")
    import pickle
    if os.path.exists("./hashing/hashing_{}.pickle".format(dataset_name)):
        with open("./hashing/hashing_{}.pickle".format(dataset_name), 'rb') as f:
            images1 = pickle.load(f)
    else:
        wb = openpyxl.load_workbook(target_xlsx)
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            image_name = str(sheet['B' + str(i)].value)
            try:
                hash = photohash.average_hash(target_folder+image_name)
                images1[target_folder+image_name] = hash
            except Exception:
                print("Hashing error {}".format(image_name))
        with open("./hashing/hashing_{}.pickle".format(dataset_name), 'wb') as f:
            pickle.dump(images1, f)

    target_dataset = images1
    num_items = len(target_dataset)
    dup_clusters, cluster_sets, dup_results = {}, set(), []
    num_repeat = 0
    ### self repeat
    if os.path.exists('D://hashing_self_repeat/{}'.format(dataset_name)):
        if os.path.exists("D://hashing_self_repeat//clusters/clusters_{}.pickle".format(dataset_name)):
            with open("D://hashing_self_repeat//clusters/clusters_{}.pickle".format(dataset_name), 'rb') as f:
                dup_clusters = pickle.load(f)
        else:
            cluster_folders = os.listdir("D://hashing_self_repeat/{}".format(dataset_name))
            print("Found {} clusters.".format(len(cluster_folders)))
            for cluster_folder in cluster_folders:
                cluster_folder = 'D://hashing_self_repeat/{}/{}'.format(dataset_name,cluster_folder)
                result_cluster = set()
                items = os.listdir(cluster_folder)
                for item in items:
                    result_cluster.add(target_folder+item)
                dup_clusters[cluster_folder] = result_cluster
            with open("D://hashing_self_repeat//clusters/clusters_{}.pickle".format(dataset_name), 'wb') as f:
                pickle.dump(dup_clusters, f)
    else:
        os.mkdir('D://hashing_self_repeat/{}'.format(dataset_name))
        ########## old fashion v1: list rumor_images images ###########
        # for item1 in target_dataset:
        #     already_saved = False
        #     item1_hash = target_dataset[item1]
        #     for item2 in target_dataset:
        #         item2_hash = target_dataset[item2]
        #         if item2 != item1 and photohash.hashes_are_similar(item1_hash, item2_hash, tolerance=1): #perceptual_hash(folder1+item1, folder2+item2):
        #             print("Found {} {}".format(item1,item2))
        #             # find the clusters respectively of item1 and item2
        #             idx_a, idx_b, set_a, set_b = -1,-1,set(),set()
        #             for idx in range(len(dup_clusters)):
        #                 query_cluster = dup_clusters[idx]
        #                 if item1 in query_cluster:
        #                     set_a,idx_a = query_cluster,idx
        #                 if item2 in query_cluster:
        #                     set_b,idx_b = query_cluster,idx
        #                 if len(set_a) !=0 and len(set_b) !=0:
        #                     break
        #             result_cluster = set_a.union(set_b)
        #             result_cluster.add(item1)
        #             result_cluster.add(item2)
        #             max_idx, min_idx = max(idx_a,idx_b), min(idx_a,idx_b)
        #             if max_idx!=-1:
        #                 del dup_clusters[max_idx]
        #             if min_idx!=-1 and min_idx!=max_idx:
        #                 del dup_clusters[min_idx]
        #             dup_clusters.append(result_cluster)
        #             item1_short = item1[item1.rfind('/')+1:]
        #             if not already_saved:
        #                 # shutil.copy(item1,'D://hashing_self_repeat/{}'.format(item1_short))
        #                 already_saved = True
        #                 num_repeat += 1
        ######### v2: only look once #################
        print("using v2: only look once.")
        cluster_no = 0
        have_checked_images = set()
        for item1 in tqdm(target_dataset):
            if item1 in have_checked_images:
                continue
            print("Results for {}".format(item1))
            item1_hash = target_dataset[item1]
            result_cluster = set()
            for item2 in target_dataset:
                if item2 in have_checked_images:
                    continue
                item2_hash = target_dataset[item2]
                if item2 != item1 and photohash.hashes_are_similar(item1_hash, item2_hash, tolerance=1): #perceptual_hash(folder1+item1, folder2+item2):
                    result_cluster.add(item1)
                    result_cluster.add(item2)
                    have_checked_images.add(item2)
            if len(result_cluster) !=0:
                cluster_no += 1
                dup_clusters[item1] = result_cluster
                os.mkdir('D://hashing_self_repeat/{}/{}'.format(dataset_name,cluster_no))
                for images in result_cluster:
                    shutil.copy(images, 'D://hashing_self_repeat/{}/{}/{}'.format(dataset_name,cluster_no, images[images.rfind('/') + 1:]))
            print("Found {}".format(len(result_cluster)))

    for item1 in tqdm(target_dataset):
        dup_data, num_dups = {}, 0
        dup_image_list = ""
        for cluster_name in dup_clusters:
            query_cluster = dup_clusters[cluster_name]
            if item1 in query_cluster:
                for items in query_cluster:
                    dup_image_list = dup_image_list + items + "[sep]"
                    num_dups += 1
                num_repeat += 1
                dup_data['cluster'] = cluster_name
                break
        dup_data['path'] = item1[item1.rfind('/')+1:]
        dup_data['num'] = num_dups
        dup_data['list'] = dup_image_list
        if num_dups>5:
            if not os.path.exists('D://hashing_self_repeat/{}/mostly_repeated'.format(dataset_name)):
                os.mkdir('D://hashing_self_repeat/{}/mostly_repeated'.format(dataset_name))
            tmp_name = dup_image_list.split('[sep]')[0]
            shutil.copy(tmp_name, 'D://hashing_self_repeat/{}/mostly_repeated/{}'.format(dataset_name,tmp_name[tmp_name.rfind('/') + 1:]))

        dup_results.append(dup_data)

    df = pd.DataFrame(dup_results)
    df.to_excel('./duplicated_images_{}.xlsx'.format(dataset_name))

    print("Repeat: {} Sum: {}".format(num_repeat,num_items))

if __name__ == '__main__':
    pass
    # get_human_labels()
    # check_similarity('E:/rumor_dataset/rumor_datasets/images/All images/',
    #                  "./dataset/rumor_dataset/all_images_chinese.xlsx",
    #                  "rumor_dataset")
    # check_similarity('E:/weibo/',
    #                  "./dataset/weibo/all_images.xlsx",
    #                  "weibo")
    # # get_image_list("Weibo_21")
    # check_similarity('E:/Weibo_21/',
    #                  "./dataset/Weibo_21/all_images.xlsx",
    #                  "Weibo_21")
    # weibo_text_to_excel()

    # get_pure_dataset()
    # copy_data()
    # reload_weibo()
    # reload_xlsxs_AAAI()
    # weibo_text_to_excel(is_train="train")
    # weibo_text_to_excel(is_train="test")
    # reload_weibo(is_train="both")
    # reload_tweet()
    # generate_ambiguity()
    # generate_more_negative_AAAI()
    # print(photohash.is_look_alike(image1, image2))

