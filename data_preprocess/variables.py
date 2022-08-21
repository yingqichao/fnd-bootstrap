import openpyxl
"""
    删除规则：
    1.关键词，
    2.长宽比失衡严重或者太小，
    3.指定图像和集合，
    4.要cv2打得开，
    5.不能是问句
    6.文字去重 
    7.图像去重
"""

forbidden_words = [
    "僵尸", "贞子","陪睡","上床",
    "轮奸","情妇","强奸","太残忍","偷小孩","傻逼","呆逼","呆子","弱智",
    "智障","尼玛","妓女","卧槽","我日","去你妈","滚你妈","狗娘养的",'你妈死了',"火葬场"
                     ]
absolute_fake_words = ["寻人", "急找孩子",] #'震惊',"帮忙扩散","帮忙转一下","帮转"
not_allowed_words = forbidden_words + absolute_fake_words
thresh_width_height = 2
not_allowed_clusters = ['./hashing_self_repeat/343']
baned_hashing = set()

rumor_root = 'E:/rumor_dataset/rumor_datasets/images/All images/'
weibo_fake_root ='E:/weibo/rumor_images/'
weibo_real_root = 'E:/weibo/nonrumor_images/'
weibo21_fake_root = 'E:/Weibo_21/rumor_images/'
weibo21_real_root = 'E:/Weibo_21/nonrumor_images/'


mixset_xlsx = "./dataset/rumor_dataset/all_images.xlsx"

def get_workbook(xlsx):
    wb = openpyxl.load_workbook(xlsx)
    sheetnames = wb.sheetnames
    sheet = wb[sheetnames[0]]
    rows = sheet.max_row
    return sheet, rows


def strQ2B(ustring):
    rstring = ""
    for uchar in ustring:
        inside_code = ord(uchar)
        if inside_code == 12288:                            # 全角空格直接转换
            inside_code = 32
        elif 65281 <= inside_code <= 65374:   				# 全角字符（除空格）根据关系转化
            inside_code -= 65248
        if inside_code<=66812:
            rstring += chr(inside_code)
    return rstring

def del_emoji(ustring):
    rstring = ""
    for uchar in ustring:
        inside_code = ord(uchar)
        if inside_code<=66812:
            rstring += chr(inside_code)
    return rstring