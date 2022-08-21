import time
import json
from tkinter import *
from PIL import Image, ImageTk
import wget
import copy
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from scp import SCPClient
from tqdm import tqdm
import datetime
import random
from tkinter.filedialog import askdirectory, askopenfilename
import shutil
import options.options as option
import photohash
import cv2
import pickle
import random
random.seed(1000)

from variables import rumor_root, weibo_fake_root, weibo_real_root, weibo21_fake_root, weibo21_real_root, \
    get_workbook

def generate_ambiguity():
    #   WE NEED TO SELECT 5000 ALIGNED IMAGE+TEXT AS WELL AS 5000 PAIRS NON-ALIGNED PAIRS
    records_list = []
    dataset_name = 'Weibo_21'
    granted_positive_times, granted_negative_times = 1, 1
    # if dataset_name == 'gossip':
    #     xlsx = "./dataset/{}/{}_{}.xlsx".format(dataset_name, dataset_name, "train_no_filt")
    # else:
    #     xlsx = "./dataset/{}/{}.xlsx".format(dataset_name, "train_datasets_WWW")
    xlsxs = [f"./dataset/{dataset_name}/origin_do_not_modify/train_datasets_Weibo21.xlsx",
             f"./dataset/{dataset_name}/origin_do_not_modify/test_datasets_Weibo21.xlsx"]
    for xlsx in xlsxs:
        num_real_count, num_fake_count = 0, 0
        sheet_rumor, rows_rumor = get_workbook(xlsx)
        valid_list_positive = {} #set([i for i in range(2,rows_rumor+1)])
        for i in range(2,rows_rumor+1):
            valid_list_positive[i] = granted_positive_times
        valid_list_negative_text = {} #set([i for i in range(2, rows_rumor + 1)])
        for i in range(2,rows_rumor+1):
            valid_list_negative_text[i] = granted_negative_times
        # valid_list_negative_image = {} #set([i for i in range(2, rows_rumor + 1)])
        # for i in range(2,rows_rumor+1):
        #     valid_list_negative_image[i] = granted_negative_times


        THRESH = 2300 if "train" in xlsx else 200
        LIST = ['F', 'C', 'E', 'X']
        # GET 10000 PAIRS

        while (num_real_count<THRESH or num_fake_count<THRESH) and len(valid_list_negative_text)>0 and len(valid_list_positive)>0:  # title label content image B C E F
            # NOTE: 1 REPRESENTS REAL NEWS
            ########### Generate positive sample ####################################
            records = {}
            # idx_ranP = random.randint(0,len(valid_list_positive)-1)
            # idx_ranN = random.randint(0, len(valid_list_negative_text)-1)

            # select an image from the dataset
            iP = random.choice(list(valid_list_positive))

            images_name = str(sheet_rumor[LIST[0] + str(iP)].value)
            label = int(sheet_rumor[LIST[1] + str(iP)].value)
            content = str(sheet_rumor[LIST[2] + str(iP)].value)
            if label==0:
                del valid_list_positive[iP]
                continue

            # check if the news selected is real news and the image/text is valid, if not, remove it from all the dicts
            full_valid_name = []
            new_images = images_name.split('|')
            for new_image in new_images:
                is_found_image = cv2.imread(f"E:/{dataset_name}/" + new_image)
                if is_found_image is not None:
                    image_width, image_height = is_found_image.shape[0], is_found_image.shape[1]
                    ## CONDUCT FILTERING OUT INVALID NEWS IF NOT NO_FILTER_OUT
                    ratio = float(image_width) / image_height
                    if ratio>3 or ratio<0.33 or image_width < 100 or image_height < 100:
                        pass
                        # news_type = "text"
                        # print("Size {} too small {} skip..".format(is_found_image.shape,new_image))
                        # num_too_small += 1
                    else:
                        full_valid_name.append(new_image)
            # collect the news as positive if both image and text satisfies
            images_name = ''
            if len(full_valid_name)!=0:
                images_name = '|'.join(full_valid_name)
            if len(images_name)!=0 and len(content)>=10:
                images_name = '|'.join(full_valid_name)
                records['content'] = content
                records['image'] = images_name
                records['label'] = 0  # using soft_label
                records['debug'] = "{} {}".format(iP, iP)
                records_list.append(records)
                records = {}
                num_real_count += 1
                print(f"Got {num_real_count} positive samples")
            del valid_list_positive[iP]


            if len(images_name)!=0 and len(valid_list_negative_text)>0 and num_fake_count<THRESH:
                label, iN, content = 0, -1, ''
                while iN == -1 and len(valid_list_negative_text)>0:
                    iN = random.choice(list(valid_list_negative_text))
                    label = int(sheet_rumor[LIST[1] + str(iN)].value)
                    content = str(sheet_rumor[LIST[2] + str(iN)].value)
                    if label==0 or iN==iP or len(content)<10:
                        del valid_list_negative_text[iN]
                        iN = -1

                if iN!=-1:
                    # collect the news as negative if both image and text satisfies
                    records['content'] = content
                    records['image'] = images_name
                    records['label'] = 1
                    records['debug'] = "{} {}".format(iN,iP)
                    records_list.append(records)
                    records = {}
                    num_fake_count += 1
                    print(f"Got {num_fake_count} negative samples")
                    del valid_list_negative_text[iN]


        print("{} {} {}".format(num_real_count,num_fake_count,len(records_list)))
    ambiguity_excel = "./dataset/{}/origin_do_not_modify/{}_{}.xlsx".format(dataset_name, dataset_name, "train_ambiguity_new")
    df = pd.DataFrame(records_list)
    df.to_excel(ambiguity_excel)





def download_images_and_load():
    results, record, error_json = [], {}, 0
    json_files = ['E:/Weibo_21/fake_release_all.json', 'E:/Weibo_21/real_release_all.json']
    folders = ['E:/Weibo_21/rumor_images/', 'E:/Weibo_21/nonrumor_images/']
    conduct_download = False

    for i in range(2):
        json_file = json_files[i]
        f = open(json_file, encoding='utf-8')
        line = f.readline()  # id,content,comments,timestamp,piclists,label,category
        while line:
            try:
                folder = folders[i]
                images_set = set(os.listdir(folder))
                item = json.loads(line)
                record = {}
                piclists = item["piclists"]
                image_name = ""
                if isinstance(piclists, float): piclists = []
                if not isinstance(piclists, list): piclists = [piclists]
                for full_image_name in piclists:
                    if full_image_name[:2] == '//': full_image_name = "http:" + full_image_name
                    short_name = full_image_name[full_image_name.rfind('/') + 1:]
                    if short_name[-4:] != 'gif':
                        if short_name not in images_set:
                            if conduct_download:
                                try:
                                    wget.download(full_image_name, folder + short_name)
                                    image_name = image_name + "rumor_images/" + short_name + "|"
                                    print("Download ok. {}".format(full_image_name))
                                except Exception:
                                    print("Download error. {}".format(full_image_name))
                            else:
                                print("Do not download. {}".format(full_image_name))
                        else:
                            image_name = image_name + "rumor_images/" + short_name + "|"
                            print("Already Downloaded. {}".format(full_image_name))
                    else:
                        print("Gif Skipped. {}".format(full_image_name))

                record['source'] = ""
                record['images'] = image_name
                record['content'] = item["content"]
                record['label'] = i
                results.append(record)
                line = f.readline()
            except Exception:
                print("Load JSON error!")
                error_json += 1
        f.close()

    df = pd.DataFrame(results)
    print(f"Num of news: {len(results)}")
    df.to_excel('./dataset/Weibo_21/origin_do_not_modify/all_datasets_Weibo21_origin.xlsx')

def reload_Weibo21():
    """
    自己根据8：1：1划分数据集
    这是曹老师那边邮件里说的划分方法
    """

    num_length, num_cv_error, num_no_valid = 0,0,0


    train_list, val_list, test_list = [],[],[]
    folder1, folder2 = "E:/Weibo_21/nonrumor_images", "E:/Weibo_21/rumor_images"
    images1, images2 = set(os.listdir(folder1)), set(os.listdir(folder2))
    images_set = images1.union(images2)
    news_xlsxs = []
    news_xlsxs.append('./dataset/Weibo_21/origin_do_not_modify/all_datasets_Weibo21_origin.xlsx')
    num_count = 0

    for idx, news_xlsx in enumerate(news_xlsxs):
        wb = openpyxl.load_workbook(news_xlsx)
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            source = sheet['B' + str(i)].value
            title, full_valid_name = '', []
            content = sheet['D' + str(i)].value
            if content is None or len(content) <=5:
                print("Found length not satisfy {}".format(content))
                num_length += 1
                continue
            image_names = sheet['C' + str(i)].value
            label = int(sheet['E' + str(i)].value)
            if image_names is not None and len(image_names) != 0:
                image_name_list = image_names.split('|')[:-1]
                for image_name in image_name_list:
                    full_image_name = image_name.lstrip(' ').rstrip(' ')  # http url
                    short_name = full_image_name[full_image_name.rfind('/') + 1:]
                    if short_name in images_set:
                        new_image = ("rumor_images/" if label == 0 else "nonrumor_images/") + short_name
                        is_found_image = cv2.imread("E:/Weibo_21/"+new_image) is not None
                        if is_found_image:
                            full_valid_name.append(new_image)
            if len(full_valid_name)!=0:
                full_valid_name = '|'.join(full_valid_name)
                record = {}
                record['title'], record['label'], record['source'], record['content'], record['image'] \
                    = title, label, source, content, full_valid_name

                if num_count%10<8:
                    train_list.append(record)
                elif num_count%10==8:
                    # pass
                    val_list.append(record)
                elif num_count%10==9:
                    test_list.append(record)

                num_count += 1

            else:
                print("No valid image found {}".format(content))
                num_no_valid += 1

    filename = './dataset/Weibo_21/origin_do_not_modify/train_datasets.xlsx'
    df = pd.DataFrame(train_list)
    df.to_excel(filename)
    print(f"File saved at {filename}")
    filename = './dataset/Weibo_21/origin_do_not_modify/val_datasets.xlsx'
    df = pd.DataFrame(val_list)
    df.to_excel(filename)
    print(f"File saved at {filename}")
    filename = './dataset/Weibo_21/origin_do_not_modify/test_datasets.xlsx'
    df = pd.DataFrame(test_list)
    df.to_excel(filename)
    print(f"File saved at {filename}")
    print("length {} CV_Error {} No_valid {}".format(num_length,num_cv_error, num_no_valid))
    print("Train_set {} Test_set {} Val_set {}".format(len(train_list),len(test_list),len(val_list)))

if __name__ == '__main__':
    generate_ambiguity()