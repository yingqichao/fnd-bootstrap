import time
import json
from tkinter import *
from PIL import Image, ImageTk
import wget
import copy
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from scp import SCPClient
from tqdm import tqdm
import datetime
import random
from tkinter.filedialog import askdirectory, askopenfilename
import shutil
import options.options as option
import photohash
import cv2
# from imagededup.methods import PHash
# phasher = PHash()
import pickle
import random
random.seed(1000)

from variables import rumor_root, weibo_fake_root, weibo_real_root, weibo21_fake_root, weibo21_real_root, \
    get_workbook, not_allowed_clusters, not_allowed_words


def get_pure_dataset():
    hashing_not_allowed, ban_images = {}, os.listdir('./ban_images')
    for ban_image in ban_images:
        hash = photohash.average_hash('./ban_images/' + ban_image)
        hashing_not_allowed[ban_image] = hash
    print("Finished loading ban images..")
    rumor_folder = 'E:/rumor_dataset/rumor_datasets/images/All images/'
    weibo_folder = 'E:/weibo/'
    weibo21_folder = 'E:/Weibo_21/'
    xlsx_rumor = './dataset/rumor_dataset/origin_do_not_modify/train_datasets_Qian.xlsx'
    xlsx_weibo = './dataset/weibo/origin_do_not_modify/train_datasets_weibo.xlsx'
    xlsx_weibo21 = './dataset/Weibo_21/origin_do_not_modify/train_datasets_Weibo21.xlsx'
    # dataset_rumor = './dataset/rumor_dataset/news_datasets.xlsx'
    # dataset_weibo = './dataset/weibo/news_datasets.xlsx'
    # dataset_weibo21 = './dataset/Weibo_21/news_datasets.xlsx'
    dups_rumor = './duplicated_images_rumor_dataset.xlsx'
    dups_weibo = './duplicated_images_weibo.xlsx'
    dups_weibo21 = './duplicated_images_Weibo_21.xlsx'
    previous_index = -1
    hashing_rumor, hashing_weibo, topics = {}, {}, {}
    record, results,status, has_recorded = {}, [], [], {}
    num_word_len, num_invalid_word, num_similar, num_invalid_hashing, num_invalid_format, num_invalid_size, num_self_repeat = 0,0,0,0,0,0,0
    with open("./hashing/hashing_rumor_dataset.pickle", 'rb') as f:
        hashing_rumor = pickle.load(f)
    with open("./hashing/hashing_weibo.pickle", 'rb') as f:
        hashing_weibo = pickle.load(f)
    with open("./hashing/hashing_weibo_21.pickle", 'rb') as f:
        hashing_weibo21 = pickle.load(f)
    # rumor xlsx workbook
    sheet_rumor, rows_rumor = get_workbook(xlsx_rumor)
    sheet_rumor_dups, rows_rumor_dups = get_workbook(dups_rumor)
    # sheet_rumor_dataset, rows_sheet_rumor = get_workbook(dataset_rumor)
    # weibo xlsx workbook
    sheet_weibo, rows_weibo = get_workbook(xlsx_weibo)
    sheet_weibo_dups, rows_weibo_dups = get_workbook(dups_weibo)
    # sheet_weibo_dataset, rows_sheet_weibo = get_workbook(dataset_weibo)
    # weibo21 xlsx workbook
    sheet_weibo21, rows_weibo21 = get_workbook(xlsx_weibo21)
    sheet_weibo21_dups, rows_weibo21_dups = get_workbook(dups_weibo21)
    # sheet_weibo21_dataset, rows_sheet_weibo21 = get_workbook(dataset_weibo21)

    # duplicate folder
    dup_query_dict_rumor = {}
    for i in tqdm(range(2, rows_rumor_dups + 1)):
        dup_query_dict_rumor[str(sheet_rumor_dups['B' + str(i)].value)] = int(sheet_rumor_dups['C' + str(i)].value)
    dup_query_dict_weibo = {}
    for i in tqdm(range(2, rows_weibo_dups + 1)):
        dup_query_dict_weibo[str(sheet_weibo_dups['B' + str(i)].value)] = int(sheet_weibo_dups['D' + str(i)].value)
    dup_query_dict_weibo21 = {}
    for i in tqdm(range(2, rows_weibo21_dups + 1)):
        dup_query_dict_weibo21[str(sheet_weibo_dups['B' + str(i)].value)] = int(sheet_weibo21_dups['C' + str(i)].value)

    the_sheets = [sheet_rumor, sheet_weibo, sheet_weibo21]
    the_rows = [rows_rumor, rows_weibo, rows_weibo21]
    self_hashings = [hashing_rumor, hashing_weibo, hashing_weibo21]
    dup_query_dicts = [dup_query_dict_rumor, dup_query_dict_weibo, dup_query_dict_weibo21]
    root_folders = [rumor_folder, weibo_folder, weibo21_folder]

    # topics
    for j, sheet in enumerate(the_sheets):
        for i in tqdm(range(2, the_rows[j]+1)):
            content = str(sheet['E' + str(i)].value)
            if content not in topics:
                topics[content] = 1
            else:
                topics[content] = topics[content]+1




    for j in range(3):

        sheet_all, rows_all = the_sheets[j], the_rows[j]
        self_hashing,dup_query_dict,root_folder = self_hashings[j], dup_query_dicts[j], root_folders[j]


        for i in tqdm(range(2, rows_all + 1)):
            # try:
            images_name = str(sheet_all['F' + str(i)].value)
            label = int(sheet_all['C' + str(i)].value)
            content = str(sheet_all['E' + str(i)].value)

            record = {}
            images = images_name.split('|')
            valid_image_name = []
            for image_name in images:
                hash_key = root_folder + (image_name if j != 0 else image_name[image_name.rfind('/') + 1:])
                if hash_key in self_hashing:
                    item1_hash = self_hashing[hash_key]
                else:
                    print(f"Not found {hash_key}")
                    continue
                ###### 1. check whether the word length is not enough
                if len(content)<5:
                    print("Word length not satisfied {} skip..".format(image_name))
                    num_word_len += 1
                    continue
                ###### 2. check whether the news contains forbiddened words
                found_invalid_word = None
                if content[-1] == '?' or content[-1] == '？':
                    found_invalid_word = "[Ends With ?]"
                else:
                    for word in not_allowed_words:
                        if word in content:
                            found_invalid_word = word
                            break
                if found_invalid_word is not None:
                    print("Invalid word found {} {}skip..".format(found_invalid_word,image_name))
                    num_invalid_word +=1
                    continue

                ###### 4. check if the image is forbidden
                found_invalid_hashing = False
                for key in hashing_not_allowed:
                    item2_hash = hashing_not_allowed[key]
                    if item1_hash is not None and photohash.hashes_are_similar(item1_hash, item2_hash, tolerance=1):
                        found_invalid_hashing = True
                        break
                if found_invalid_hashing:
                    print("Invalid image found {} skip..".format(image_name))
                    num_invalid_hashing += 1
                    continue
                ###### 5. check similarity with weibo (only outside weibo)
                if j!=1:
                    found_similar = False
                    for key in hashing_weibo:
                        item2_hash = hashing_weibo[key]
                        if photohash.hashes_are_similar(item1_hash, item2_hash, tolerance=0.05):
                            found_similar = True
                            break
                    if found_similar:
                        print("Similar item found {} skip..".format(image_name))
                        num_similar += 1
                        continue
                ###### 6. check whether the file can be open by cv2

                image_open = cv2.imread(hash_key)
                if image_open is None:
                    print("Cannot open {} skip..".format(image_name))
                    num_invalid_format += 1
                    continue
                ###### 7. check the ratio of the image
                image_width, image_height = image_open.shape[0], image_open.shape[1]
                ratio = float(image_open.shape[0])/image_open.shape[1]
                if ratio>3 or ratio<0.33 or image_width<100 or image_height<100:
                    print("File size not valid {} {} {}".format(image_width, image_height,image_name))
                    num_invalid_size += 1
                    continue
                ###### 3. check if the image is highly self-repeated
                times = (float(dup_query_dict[hash_key]) + 1) if hash_key in dup_query_dict else 1
                possibly1 = (5 if label == 1 else 5) / times
                possibly2 = 1  # (1.5 if label==1 else 5)/(float(topics[content]))
                if np.random.random() >= possibly1 or np.random.random() >= possibly2:
                    print("Self repeat deletion image {} topic {} {} skip.."
                          .format(times, topics[content], image_name))
                    num_self_repeat += 1
                    continue

                ##### if rumor_images the above is valid
                valid_image_name.append(image_name)

            if len(valid_image_name)!=0:
                image_name = '|'.join(valid_image_name)
                record['title'], record['label'], record['source'], record['content'], record['image'] \
                    = '',label,'',content,image_name
                results.append(record)
            # else:
            #     image_prelist = record['image']
            #     record['image'] = "{}|{}".format(image_prelist,image_name)
            # except Exception:
            #     print("Unknown error..skip..")

        status.append([len(results),num_invalid_word, num_similar, num_invalid_hashing, num_invalid_format, num_invalid_size,
        num_self_repeat, num_word_len])

    df = pd.DataFrame(results)
    df.to_excel('./cleaned_data.xlsx')
    print("records {} word {} dup {} forbid {} format {} size {} repeat {} wordlen {}")
    for item in status:
        print(item)



def copy_data():
    folder_real, folder_fake = 'E:/MixSet/nonrumor_images/', 'E:/MixSet/rumor_images/'
    xlsx_rumor = './cleaned_data.xlsx'
    sheet_rumor, rows_rumor = get_workbook(xlsx_rumor)
    rumor_imgs = set(os.listdir(rumor_root))
    weibo_fake_imgs = set(os.listdir(weibo_fake_root))
    weibo_real_imgs = set(os.listdir(weibo_real_root))
    weibo21_fake_imgs = set(os.listdir(weibo21_fake_root))
    weibo21_real_imgs = set(os.listdir(weibo21_real_root))
    for i in tqdm(range(2,rows_rumor+1)):# title label content image B C E F

        images_name = str(sheet_rumor['F' + str(i)].value)
        label = int(sheet_rumor['C' + str(i)].value)
        content = str(sheet_rumor['E' + str(i)].value)
        imgs = images_name.split('|')
        for img in imgs:
            if '/' in img:
                img = img[img.rfind('/')+1:]

            if label==0:
                if img in weibo_fake_imgs:
                    if not os.path.exists(folder_fake+img):
                        shutil.copy(weibo_fake_root+img, folder_fake+img)
                        print("copied {}".format(img))
                elif img in weibo21_fake_imgs:
                    if not os.path.exists(folder_fake + img):
                        shutil.copy(weibo21_fake_root+img, folder_fake+img)
                        print("copied {}".format(img))
                elif img in rumor_imgs:
                    if not os.path.exists(folder_fake + img):
                        shutil.copy(rumor_root + img, folder_fake + img)
                        print("copied {}".format(img))
                else:
                    print("Not found {}".format(img))
            else:
                if img in weibo_real_imgs:
                    if not os.path.exists(folder_real + img):
                        shutil.copy(weibo_real_root+img, folder_real+img)
                        print("copied {}".format(img))
                elif img in weibo21_real_imgs:
                    if not os.path.exists(folder_real + img):
                        shutil.copy(weibo21_real_root+img, folder_real+img)
                        print("copied {}".format(img))
                elif img in rumor_imgs:
                    if not os.path.exists(folder_real + img):
                        shutil.copy(rumor_root+img, folder_real+img)
                        print("copied {}".format(img))
                else:
                    print("Not found {}".format(img))


def get_human_labels():
    # wb = openpyxl.load_workbook('./labels_image.xlsx')
    # sheetnames = wb.sheetnames
    # sheet = wb[sheetnames[0]]
    # rows = sheet.max_row
    # delete_path = set()
    # for i in tqdm(range(2, rows + 1)):
    #     path = sheet['B' + str(i)].value
    #     image = cv2.imread("E:/rumor_dataset/rumor_datasets/images/All images/{}".format(path))
    #     if image is None:
    #         print(path)
    #         delete_path.add(path)

    results_dict = {}
    wb = openpyxl.load_workbook('./dataset/rumor_dataset/all_images.xlsx')
    sheetnames = wb.sheetnames
    sheet = wb[sheetnames[0]]
    rows = sheet.max_row
    for i in tqdm(range(2, rows + 1)):
        init_dict = {}
        # begin a new record
        image_path = sheet['B' + str(i)].value
        if cv2.imread(rumor_root+image_path) is not None:
            init_dict['path'] = image_path

            init_dict['label'] = float(sheet['C' + str(i)].value)
            init_dict['num'] = 0.0
            init_dict['pred_label'] = 0.0
            results_dict[sheet['B' + str(i)].value] = init_dict
        else:
            print("{} is invalid".format(image_path))

    xlsx_files = os.listdir('./results/rumor_dataset/')
    for xlsx1 in xlsx_files:
        xlsx1 = './results/rumor_dataset/' + xlsx1
        print("Collecting datas from {}".format(xlsx1))
        wb = openpyxl.load_workbook(xlsx1)
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        for i in tqdm(range(2, rows + 1)):
            img_path = sheet['B' + str(i)].value
            if img_path in results_dict:
                init_dict = results_dict[img_path]
                # init_dict['path'] = sheet['B' + str(i)].value
                # init_dict['label'] = int(sheet['C' + str(i)].value)
                prev_pred_label = init_dict['pred_label']
                prev_num = init_dict['num']
                init_dict['pred_label'] = (prev_pred_label*prev_num+float(sheet['D' + str(i)].value))/(prev_num+1)
                init_dict['num'] = prev_num + 1
                # init_dict['consistency'] = int(sheet['E' + str(i)].value)
                results_dict[sheet['B' + str(i)].value] = init_dict
            else:
                print("Not found:{}".format(img_path))

    results = []
    for items in results_dict:
        record = results_dict[items]
        record['consistency'] = 1.0 if round(record['pred_label'])==record['label'] else 0.0
        results.append(record)

    df = pd.DataFrame(results)
    df.to_excel('./labels_image.xlsx')

if __name__ == '__main__':
    get_pure_dataset()