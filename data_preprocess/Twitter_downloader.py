import copy
import os

import openpyxl
import pandas as pd
from tqdm import tqdm
import shutil
not_found_set = set()
num_rumor, num_Twitter, num_not_found = 0,0,0
results, record, error_json = [], {}, 0
json_files = ['E:/Twitter/image-verification-corpus-master/mediaeval2016/devset/posts.txt']
folders = ['E:/Twitter/image-verification-corpus-master/mediaeval2015/devset/MediaEval2015_DevSet_Images/Medieval2015_DevSet_Images/']
real_images_dict, fake_images_dict = {}, {}
non_dup_entry = set()

import cv2

def copy_image(src, dst, content, filter=False):
    if filter:
        ###### 1. check whether the word length is not enough
        if len(content) < 5:
            print("Word length not satisfied {} skip..".format(image_name))
            return 0
        ###### 6. check whether the file can be open by cv2
        image_open = cv2.imread(src)
        if image_open is None:
            print("Cannot open {} skip..".format(image_name))
            return 0
        ###### 7. check the ratio of the image
        image_width, image_height = image_open.shape[0], image_open.shape[1]
        ratio = float(image_open.shape[0]) / image_open.shape[1]
        if ratio > 2 or ratio < 0.5 or image_width < 100 or image_height < 100:
            print("File size not valid {} {} {}".format(image_width, image_height, image_name))
            return 0

    shutil.copy(src,dst)
    return 1

### part 1: Twitter dataset
for i in tqdm(range(len(json_files))):
    json_file = json_files[i]
    folder = folders[i]
    sub_folders = os.listdir(folder)
    for sub_folder in sub_folders:
        real_folder = folder + sub_folder + '/reals/'
        fake_folder = folder + sub_folder + '/fakes/'
        real_imgs = os.listdir(real_folder) if os.path.exists(real_folder) else []
        fake_imgs = os.listdir(fake_folder) if os.path.exists(fake_folder) else []
        for real_img in real_imgs:
            real_images_dict[real_img[:real_img.rfind('.')]] = real_folder+real_img
        for fake_img in fake_imgs:
            fake_images_dict[fake_img[:fake_img.rfind('.')]] = fake_folder+fake_img

    # download if not found
    url_dict = {}
    f = open('E:/Downloads/image-verification-corpus-master/image-verification-corpus-master/set_images.txt', encoding='utf-8')
    line = f.readline()  # id,content,comments,timestamp,piclists,label,category
    line_num = 0
    while line:
        if line_num != 0: # boston_fake_35	http://imgur.com/a/sUrnA	fake	boston
            item = line.split('\t')
            image_name = item[0]
            image_url = item[1]
            # label = item[2]
            url_dict[image_name] = image_url
        line = f.readline()
        line_num += 1
    f.close()

    f = open(json_file,encoding='utf-8')
    line = f.readline()  # id,content,comments,timestamp,piclists,label,category
    line_num = 0
    while line:
        if line_num!=0:
            # tweetId0 tweetText1 userId2 imageId(s)3 username4 timestamp5 label6
            item = line.split('\t')
            record = {}
            record['source'] = ""
            record['content'] = item[1]
            record['label'] = 0 if 'fake' in item[6] else 1
            imgs = item[3]
            image_name = ""
            # if isinstance(piclists,float): piclists = []
            # if not isinstance(piclists,list): piclists = [piclists]
            piclists = imgs.split(',')
            for full_image_name in piclists:
                target_dict = real_images_dict if record['label']==1 else fake_images_dict
                if full_image_name not in target_dict:
                    # try downloading
                    # try:
                    if full_image_name in url_dict:
                        url = url_dict[full_image_name]
                        if '.' in url[url.rfind('/')+1:]:
                            import wget
                            url = url[:-1] if '\n' in url else url
                            target_path = 'E:/Twitter/image-verification-corpus-master/mediaeval2015/devset/MediaEval2015_DevSet_Images/Medieval2015_DevSet_Images/Downloads/{}/{}'.format("reals" if record['label']==1 else "fakes", url[url.rfind('/')+1:])
                            print("Downloading {}".format(url))
                            wget.download(url, target_path)
                            target_dict[full_image_name] = target_path
                            print("Download ok. {}".format(full_image_name))
                    else:
                        print("Not found. {}".format(full_image_name))

                if full_image_name in target_dict:
                    image_name = image_name+ full_image_name+"|"
                    sub_name = 'nonrumor_images' if record['label']==1 else 'rumor_images'
                    pic_name_idx = target_dict[full_image_name].rfind('/')
                    pic_name = target_dict[full_image_name][pic_name_idx+1:]
                    copy_image(target_dict[full_image_name],
                                'E:/MixSet/english/{}/{}'.format(sub_name,pic_name),
                               record['content'])
                else:
                    not_found_set.add(full_image_name)

            if len(image_name)!=0:
                record['images'] = image_name[:-1]
                non_dup_entry.add(item[1])
                results.append(record)
                num_Twitter +=1
            else:
                print("Found empty entry {}".format(imgs))
                num_not_found += 1

        line = f.readline()
        line_num+=1
    f.close()

### part 2: rumor dataset
folder = "E:/rumor_dataset/rumor_datasets/images/All images/"
images = os.listdir(folder)
images_set = set(images)
news_xlsx = './dataset/rumor_dataset/news_datasets.xlsx'
wb = openpyxl.load_workbook(news_xlsx)
sheetnames = wb.sheetnames
sheet = wb[sheetnames[0]]
rows = sheet.max_row
for i in range(2, rows + 1):
    if ((i < 1001) or (i > 5001 and i < 6002)): #English
        image_names = sheet['E' + str(i)].value
        if len(image_names) != 0:
            image_name_list = set()
            image_name_list_1 = image_names.split('\t')
            image_name_list_2 = image_names.split(';')
            if len(image_name_list_1) >= len(image_name_list_2):
                image_name_list = set(image_name_list_1)
            else:
                image_name_list = set(image_name_list_2)
            for image_name in image_name_list:
                image_name = image_name.lstrip(' ').rstrip(' ')
                record = {}
                if image_name in images_set:
                    record['images'] = copy.copy(image_name)
                    record['label'] = int(sheet['F' + str(i)].value)
                    record['source'] = int(sheet['F' + str(i)].value)
                    record['content'] = int(sheet['F' + str(i)].value)
                    record['index'] = i
                    if record['content'] not in non_dup_entry:
                        results.append(record)
                        sub_name = 'nonrumor_images' if record['label']==1 else 'rumor_images'
                        num_rumor += copy_image(folder+image_name,
                                    'E:/MixSet/english/{}/{}'.format(sub_name, image_name),
                                   record['content'])
                    else:
                        print("Duplicated entry {}".format(record['content']))

df = pd.DataFrame(results)
df.to_excel('./cleaned_data_english.xlsx')
print("Twitter {} Rumor {} NotFound {} MissingImages {}"
      .format(num_Twitter,num_rumor,num_not_found, len(not_found_set)))
